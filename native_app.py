#!/usr/bin/env python3
from __future__ import annotations

import json
import logging
import os
import queue
import subprocess
import sys
import threading
import time
import traceback
from pathlib import Path
from tkinter import BooleanVar, StringVar, Tk, Toplevel, filedialog, messagebox, ttk

from openpyxl import load_workbook

from app import workbook_info_payload
from auth_store import AuthStore
from workbook import BatchOptions, process_batch


APP_NAME = "小红书笔记截图工具"


def app_data_dir() -> Path:
    if sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    elif os.name == "nt":
        base = Path(os.environ.get("LOCALAPPDATA") or Path.home() / "AppData" / "Local")
    else:
        base = Path(os.environ.get("XDG_DATA_HOME") or Path.home() / ".local" / "share")
    path = base / "xhs-screenshot-tool"
    path.mkdir(parents=True, exist_ok=True)
    return path


def setup_logging() -> None:
    log_path = app_data_dir() / "app.log"
    logging.basicConfig(
        filename=log_path,
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )


def new_native_run_dir() -> Path:
    path = app_data_dir() / "runs" / f"{time.strftime('%Y%m%d-%H%M%S')}-{os.getpid()}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def account_store() -> AuthStore:
    return AuthStore(app_data_dir() / "accounts.json")


def int_value(values: dict, key: str, default: int) -> int:
    raw = str(values.get(key) or "").strip()
    return int(raw) if raw else default


def make_batch_options(workbook_path: Path, run_dir: Path, values: dict) -> BatchOptions:
    end_row = str(values.get("end_row") or "").strip()
    return BatchOptions(
        workbook_path=workbook_path,
        run_dir=run_dir,
        sheet_name=str(values.get("sheet_name") or "Sheet1"),
        link_col=str(values.get("link_col") or "A"),
        output_col=str(values.get("output_col") or "B"),
        start_row=int_value(values, "start_row", 2),
        end_row=int(end_row) if end_row else None,
        resume=bool(values.get("resume", True)),
        timeout_ms=int_value(values, "timeout_ms", 25000),
        min_delay_ms=int_value(values, "min_delay_ms", 3500),
        max_delay_ms=int_value(values, "max_delay_ms", 60000),
        max_retries=int_value(values, "max_retries", 2),
        max_consecutive_failures=int_value(values, "max_consecutive_failures", 5),
    )


class ScreenshotApp:
    def __init__(self) -> None:
        self.auth = account_store()
        self.current_user: dict | None = None
        self.root = Tk()
        self.root.title(APP_NAME)
        self.root.geometry("880x680")
        self.root.minsize(760, 560)
        self.root.after(300, self.bring_to_front)
        self.selected_file: Path | None = None
        self.current_run_dir: Path | None = None
        self.worker_queue: queue.Queue[tuple[str, object]] = queue.Queue()

        self.sheet_name = StringVar(value="Sheet1")
        self.link_col = StringVar(value="A")
        self.output_col = StringVar(value="B")
        self.start_row = StringVar(value="2")
        self.end_row = StringVar(value="")
        self.timeout_ms = StringVar(value="25000")
        self.min_delay_ms = StringVar(value="3500")
        self.max_delay_ms = StringVar(value="60000")
        self.max_retries = StringVar(value="2")
        self.max_consecutive_failures = StringVar(value="5")
        self.resume = BooleanVar(value=True)
        self.status = StringVar(value="请选择 Excel 文件。")
        self.progress = StringVar(value="等待开始")
        self.output_path: Path | None = None
        self.columns_by_sheet: dict[str, list[dict]] = {}
        self.batch_started_at = 0.0

        self.root.withdraw()
        self.show_login()

    def bring_to_front(self) -> None:
        self.root.lift()
        self.root.focus_force()
        self.root.attributes("-topmost", True)
        self.root.after(700, lambda: self.root.attributes("-topmost", False))

    def show_login(self) -> None:
        if not self.auth.has_accounts():
            self.show_first_admin_setup()
            return
        win = Toplevel(self.root)
        win.title("登录")
        win.geometry("360x240")
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", self.root.destroy)
        win.grab_set()

        frame = ttk.Frame(win, padding=22)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="登录小红书笔记截图工具", font=("", 16, "bold")).pack(anchor="w", pady=(0, 12))
        username = StringVar()
        password = StringVar()
        ttk.Label(frame, text="账号").pack(anchor="w")
        ttk.Entry(frame, textvariable=username).pack(fill="x", pady=(3, 10))
        ttk.Label(frame, text="密码").pack(anchor="w")
        password_entry = ttk.Entry(frame, textvariable=password, show="•")
        password_entry.pack(fill="x", pady=(3, 14))

        def submit() -> None:
            user = self.auth.verify_login(username.get(), password.get())
            if not user:
                messagebox.showerror(APP_NAME, "账号或密码错误，或账号已被禁用。")
                return
            self.current_user = user
            win.destroy()
            self.open_main_window()

        ttk.Button(frame, text="登录", command=submit).pack(fill="x")
        win.bind("<Return>", lambda _event: submit())
        username.set("admin")
        password_entry.focus_force()
        self.center_window(win)

    def show_first_admin_setup(self) -> None:
        win = Toplevel(self.root)
        win.title("首次创建管理员")
        win.geometry("390x270")
        win.resizable(False, False)
        win.protocol("WM_DELETE_WINDOW", self.root.destroy)
        win.grab_set()

        frame = ttk.Frame(win, padding=22)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="首次使用：创建管理员账号", font=("", 16, "bold")).pack(anchor="w", pady=(0, 8))
        ttk.Label(frame, text="这个管理员用于管理本软件的登录账号。", foreground="#667085").pack(anchor="w", pady=(0, 12))
        username = StringVar(value="admin")
        password = StringVar()
        confirm = StringVar()
        ttk.Label(frame, text="管理员账号").pack(anchor="w")
        ttk.Entry(frame, textvariable=username).pack(fill="x", pady=(3, 8))
        ttk.Label(frame, text="密码（至少 6 位）").pack(anchor="w")
        ttk.Entry(frame, textvariable=password, show="•").pack(fill="x", pady=(3, 8))
        ttk.Label(frame, text="确认密码").pack(anchor="w")
        ttk.Entry(frame, textvariable=confirm, show="•").pack(fill="x", pady=(3, 12))

        def submit() -> None:
            if password.get() != confirm.get():
                messagebox.showerror(APP_NAME, "两次密码不一致。")
                return
            try:
                self.current_user = self.auth.create_first_admin(username.get(), password.get())
            except Exception as exc:
                messagebox.showerror(APP_NAME, str(exc))
                return
            win.destroy()
            self.open_main_window()

        ttk.Button(frame, text="创建并进入", command=submit).pack(fill="x")
        win.bind("<Return>", lambda _event: submit())
        self.center_window(win)

    def open_main_window(self) -> None:
        self.root.deiconify()
        self.build_ui()
        self.root.after(300, self.bring_to_front)
        self.root.after(500, self.refresh_today_logs)

    def center_window(self, win: Toplevel) -> None:
        win.update_idletasks()
        width = win.winfo_width()
        height = win.winfo_height()
        x = (win.winfo_screenwidth() - width) // 2
        y = (win.winfo_screenheight() - height) // 2
        win.geometry(f"{width}x{height}+{x}+{y}")
        win.lift()
        win.focus_force()

    def build_ui(self) -> None:
        outer = ttk.Frame(self.root, padding=18)
        outer.pack(fill="both", expand=True)

        title = ttk.Label(outer, text=APP_NAME, font=("", 22, "bold"))
        title.pack(anchor="w")
        note = ttk.Label(
            outer,
            text="在这个窗口里选择 Excel、工作表、链接列和输出列；截图结果会写回新的 Excel。",
            foreground="#667085",
        )
        note.pack(anchor="w", pady=(4, 14))
        user_text = f"当前账号：{self.current_user.get('username', '') if self.current_user else ''}"
        if self.current_user and self.current_user.get("role") == "admin":
            user_row = ttk.Frame(outer)
            user_row.pack(fill="x", pady=(0, 12))
            ttk.Label(user_row, text=user_text, foreground="#667085").pack(side="left")
            ttk.Button(user_row, text="账号管理中心", command=self.open_account_center).pack(side="right")
        else:
            ttk.Label(outer, text=user_text, foreground="#667085").pack(anchor="w", pady=(0, 12))

        file_box = ttk.LabelFrame(outer, text="Excel 文件")
        file_box.pack(fill="x", pady=(0, 12))
        row = ttk.Frame(file_box, padding=10)
        row.pack(fill="x")
        self.file_label = ttk.Label(row, text="尚未选择文件")
        self.file_label.pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="选择 Excel", command=self.choose_file).pack(side="right")

        settings = ttk.LabelFrame(outer, text="批量设置")
        settings.pack(fill="x", pady=(0, 12))
        grid = ttk.Frame(settings, padding=10)
        grid.pack(fill="x")
        for index in range(4):
            grid.columnconfigure(index, weight=1)

        self.sheet_combo = self.add_combo(grid, "工作表", self.sheet_name, 0, 0)
        self.link_combo = self.add_combo(grid, "链接列", self.link_col, 0, 1)
        self.output_combo = self.add_combo(grid, "输出列", self.output_col, 0, 2)
        self.add_entry(grid, "起始行", self.start_row, 0, 3)
        self.add_entry(grid, "结束行（可空）", self.end_row, 2, 0)
        self.add_entry(grid, "单页超时(ms)", self.timeout_ms, 2, 1)
        self.add_entry(grid, "基础间隔(ms)", self.min_delay_ms, 2, 2)
        self.add_entry(grid, "最大退避(ms)", self.max_delay_ms, 2, 3)
        self.add_entry(grid, "重试次数", self.max_retries, 4, 0)
        self.add_entry(grid, "连续失败中止", self.max_consecutive_failures, 4, 1)
        ttk.Checkbutton(grid, text="跳过已标记成功的行", variable=self.resume).grid(row=5, column=0, columnspan=2, sticky="w", pady=(10, 0))

        self.sheet_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_columns())

        actions = ttk.Frame(outer)
        actions.pack(fill="x", pady=(0, 12))
        self.start_button = ttk.Button(actions, text="开始批量截图", command=self.start_batch)
        self.start_button.pack(side="left")
        self.open_output_button = ttk.Button(actions, text="打开结果位置", command=self.open_output, state="disabled")
        self.open_output_button.pack(side="left", padx=(10, 0))

        progress_box = ttk.LabelFrame(outer, text="截图进度")
        progress_box.pack(fill="x", pady=(0, 12))
        inner = ttk.Frame(progress_box, padding=10)
        inner.pack(fill="x")
        ttk.Label(inner, textvariable=self.status, foreground="#b42318").pack(anchor="w")
        self.progress_bar = ttk.Progressbar(inner, mode="determinate", maximum=100)
        self.progress_bar.pack(fill="x", pady=(10, 6))
        ttk.Label(inner, textvariable=self.progress).pack(anchor="w")
        self.recent_failures = ttk.Treeview(inner, columns=("row", "reason"), show="headings", height=3)
        self.recent_failures.heading("row", text="行")
        self.recent_failures.heading("reason", text="最近失败原因")
        self.recent_failures.column("row", width=60, stretch=False)
        self.recent_failures.column("reason", width=720)
        self.recent_failures.pack(fill="x", pady=(8, 0))

        log_box = ttk.LabelFrame(outer, text="当天截图日志")
        log_box.pack(fill="both", expand=True)
        log_frame = ttk.Frame(log_box, padding=10)
        log_frame.pack(fill="both", expand=True)
        self.log_text = ttk.Treeview(log_frame, columns=("time", "summary"), show="headings", height=8)
        self.log_text.heading("time", text="时间")
        self.log_text.heading("summary", text="结果")
        self.log_text.column("time", width=90, stretch=False)
        self.log_text.column("summary", width=660)
        self.log_text.pack(fill="both", expand=True)
        ttk.Button(log_frame, text="刷新日志", command=self.refresh_today_logs).pack(anchor="e", pady=(8, 0))

    def add_combo(self, parent, label: str, variable: StringVar, row: int, col: int):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4)
        combo = ttk.Combobox(parent, textvariable=variable, state="readonly")
        combo.grid(row=row + 1, column=col, sticky="ew", padx=4, pady=(3, 10))
        return combo

    def add_entry(self, parent, label: str, variable: StringVar, row: int, col: int):
        ttk.Label(parent, text=label).grid(row=row, column=col, sticky="w", padx=4)
        entry = ttk.Entry(parent, textvariable=variable)
        entry.grid(row=row + 1, column=col, sticky="ew", padx=4, pady=(3, 10))
        return entry

    def choose_file(self) -> None:
        filename = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if filename:
            self.load_workbook(Path(filename))

    def load_workbook(self, path: Path) -> None:
        self.selected_file = path
        self.file_label.config(text=str(path))
        try:
            payload = workbook_info_payload(path)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"读取 Excel 失败：{exc}")
            return
        self.columns_by_sheet = {sheet["name"]: sheet["columns"] for sheet in payload.get("sheets", [])}
        sheet_names = list(self.columns_by_sheet)
        self.sheet_combo.config(values=sheet_names)
        recommended = payload.get("recommended", {})
        self.sheet_name.set(recommended.get("sheetName") or (sheet_names[0] if sheet_names else "Sheet1"))
        self.refresh_columns(recommended)
        self.status.set("Excel 已读取，请确认工作表和列后开始。")

    def refresh_columns(self, recommended: dict | None = None) -> None:
        recommended = recommended or {}
        columns = self.columns_by_sheet.get(self.sheet_name.get()) or [{"letter": "A", "label": "A"}, {"letter": "B", "label": "B"}]
        labels = [col["label"] for col in columns]
        self.link_combo.config(values=labels)
        self.output_combo.config(values=labels)
        self.link_col.set(self.column_label_for_letter(columns, recommended.get("linkCol") or "A"))
        self.output_col.set(self.column_label_for_letter(columns, recommended.get("outputCol") or "B"))

    def column_label_for_letter(self, columns: list[dict], letter: str) -> str:
        for col in columns:
            if col["letter"] == letter:
                return col["label"]
        return columns[0]["label"] if columns else letter

    def selected_column_letter(self, label: str) -> str:
        return label.split(" - ", 1)[0].strip()

    def current_values(self) -> dict:
        return {
            "sheet_name": self.sheet_name.get(),
            "link_col": self.selected_column_letter(self.link_col.get()),
            "output_col": self.selected_column_letter(self.output_col.get()),
            "start_row": self.start_row.get(),
            "end_row": self.end_row.get(),
            "resume": self.resume.get(),
            "timeout_ms": self.timeout_ms.get(),
            "min_delay_ms": self.min_delay_ms.get(),
            "max_delay_ms": self.max_delay_ms.get(),
            "max_retries": self.max_retries.get(),
            "max_consecutive_failures": self.max_consecutive_failures.get(),
        }

    def start_batch(self) -> None:
        if not self.selected_file:
            messagebox.showwarning(APP_NAME, "请先选择 Excel 文件。")
            return
        try:
            run_dir = new_native_run_dir()
            opts = make_batch_options(self.selected_file, run_dir, self.current_values())
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"批量设置有误：{exc}")
            return

        self.current_run_dir = run_dir
        self.batch_started_at = time.time()
        self.output_path = None
        self.open_output_button.config(state="disabled")
        self.start_button.config(state="disabled")
        self.progress_bar["value"] = 0
        self.status.set("正在截图，请保持网络稳定，不要频繁手动操作浏览器。")
        self.progress.set("准备中...")

        thread = threading.Thread(target=self.run_worker_thread, args=(opts,), daemon=True)
        thread.start()
        self.root.after(800, self.poll_worker)

    def run_worker_thread(self, opts: BatchOptions) -> None:
        try:
            result = process_batch(opts)
            self.worker_queue.put(("done", result))
        except Exception as exc:
            traceback.print_exc()
            self.worker_queue.put(("error", str(exc)))

    def poll_worker(self) -> None:
        self.update_progress_from_files()
        try:
            event, payload = self.worker_queue.get_nowait()
        except queue.Empty:
            self.root.after(1000, self.poll_worker)
            return

        self.start_button.config(state="normal")
        if event == "done":
            result = payload if isinstance(payload, dict) else {}
            workbook = result.get("workbook")
            self.output_path = Path(workbook) if workbook else None
            if self.output_path:
                self.open_output_button.config(state="normal")
            self.status.set("完成：已生成写回图片的 Excel。")
            self.update_progress_from_result(result)
            self.refresh_today_logs()
        else:
            self.status.set(f"出错：{payload}")
            self.refresh_today_logs()

    def update_progress_from_files(self) -> None:
        if not self.current_run_dir:
            return
        result_path = self.current_run_dir / "worker_results.json"
        job_path = self.current_run_dir / "worker_job.json"
        status_path = self.current_run_dir / "worker_status.json"
        results = []
        total = 0
        stopped = False
        reason = ""
        live_message = ""
        if job_path.exists():
            try:
                total = len(json.loads(job_path.read_text(encoding="utf-8")).get("items", []))
            except Exception:
                pass
        if result_path.exists():
            try:
                data = json.loads(result_path.read_text(encoding="utf-8"))
                results = data.get("results", [])
                stopped = bool(data.get("stopped"))
                reason = data.get("reason", "")
            except Exception:
                pass
        if status_path.exists():
            try:
                live_message = json.loads(status_path.read_text(encoding="utf-8")).get("message", "")
            except Exception:
                live_message = ""
        success = sum(1 for item in results if item.get("status") == "成功")
        failed = len(results) - success
        self.render_recent_failures(results)
        percent = round(len(results) / total * 100) if total else 0
        self.progress_bar["value"] = percent
        stop_text = f"；已中止：{reason}" if stopped else ""
        prefix = f"{live_message}；" if live_message else ""
        if not results and self.batch_started_at and time.time() - self.batch_started_at > 90:
            prefix = (prefix or "第一行仍在等待页面加载/浏览器响应；")
        self.progress.set(f"{prefix}{len(results)}/{total or '?'}，成功 {success}，失败 {failed}{stop_text}")

    def render_recent_failures(self, results: list[dict]) -> None:
        if not hasattr(self, "recent_failures"):
            return
        for item_id in self.recent_failures.get_children():
            self.recent_failures.delete(item_id)
        failures = [item for item in results if item.get("status") and item.get("status") != "成功"][-3:]
        for item in failures:
            reason = str(item.get("reason", ""))[:180]
            self.recent_failures.insert("", "end", values=(item.get("row", ""), reason))

    def update_progress_from_result(self, result: dict) -> None:
        results = result.get("results", [])
        success = sum(1 for item in results if item.get("status") == "成功")
        failed = len(results) - success
        self.render_recent_failures(results)
        total = len(results)
        self.progress_bar["value"] = 100 if total else 0
        self.progress.set(f"{total}/{total}，成功 {success}，失败 {failed}")

    def refresh_today_logs(self) -> None:
        for item in self.log_text.get_children():
            self.log_text.delete(item)
        runs_dir = app_data_dir() / "runs"
        today = time.strftime("%Y%m%d")
        if not runs_dir.exists():
            return
        for run_dir in sorted(runs_dir.iterdir(), reverse=True):
            if not run_dir.is_dir() or not run_dir.name.startswith(today):
                continue
            result_path = run_dir / "worker_results.json"
            results = []
            stopped = False
            reason = ""
            if result_path.exists():
                try:
                    data = json.loads(result_path.read_text(encoding="utf-8"))
                    results = data.get("results", [])
                    stopped = bool(data.get("stopped"))
                    reason = data.get("reason", "")
                except Exception:
                    pass
            success = sum(1 for item in results if item.get("status") == "成功")
            failed = len(results) - success
            time_text = run_dir.name[9:17].replace("-", "")
            if len(time_text) >= 6:
                time_text = f"{time_text[0:2]}:{time_text[2:4]}:{time_text[4:6]}"
            summary = f"共 {len(results)} 行，成功 {success}，失败 {failed}"
            if stopped:
                summary += f"，已中止：{reason}"
            self.log_text.insert("", "end", values=(time_text, summary))

    def open_output(self) -> None:
        if not self.output_path:
            return
        target = self.output_path.parent
        try:
            if sys.platform == "darwin":
                subprocess.run(["open", str(target)], check=False)
            elif os.name == "nt":
                os.startfile(str(target))  # type: ignore[attr-defined]
            else:
                subprocess.run(["xdg-open", str(target)], check=False)
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"无法打开结果位置：{exc}")

    def open_account_center(self) -> None:
        if not self.current_user or self.current_user.get("role") != "admin":
            messagebox.showwarning(APP_NAME, "只有管理员可以管理账号。")
            return
        win = Toplevel(self.root)
        win.title("账号管理中心")
        win.geometry("680x420")
        frame = ttk.Frame(win, padding=14)
        frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(frame, columns=("username", "role", "active", "updated"), show="headings", height=10)
        tree.heading("username", text="账号")
        tree.heading("role", text="角色")
        tree.heading("active", text="状态")
        tree.heading("updated", text="更新时间")
        tree.column("username", width=170)
        tree.column("role", width=90)
        tree.column("active", width=90)
        tree.column("updated", width=180)
        tree.pack(fill="both", expand=True)

        def refresh() -> None:
            for item in tree.get_children():
                tree.delete(item)
            for account in self.auth.list_accounts():
                active_text = "启用" if account["active"] else "禁用"
                tree.insert("", "end", iid=account["username"], values=(account["username"], account["role"], active_text, account["updatedAt"]))

        def selected_username() -> str | None:
            selected = tree.selection()
            return selected[0] if selected else None

        def add_or_reset() -> None:
            self.open_account_editor(refresh)

        def disable_selected() -> None:
            username = selected_username()
            if not username:
                messagebox.showwarning(APP_NAME, "请先选择账号。")
                return
            if username == self.current_user.get("username"):
                messagebox.showwarning(APP_NAME, "不能禁用当前登录账号。")
                return
            self.auth.set_active(username, False)
            refresh()

        def enable_selected() -> None:
            username = selected_username()
            if not username:
                messagebox.showwarning(APP_NAME, "请先选择账号。")
                return
            self.auth.set_active(username, True)
            refresh()

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(10, 0))
        ttk.Button(buttons, text="新增/重置账号", command=add_or_reset).pack(side="left")
        ttk.Button(buttons, text="启用", command=enable_selected).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="禁用", command=disable_selected).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="关闭", command=win.destroy).pack(side="right")
        refresh()
        self.center_window(win)

    def open_account_editor(self, on_saved) -> None:
        win = Toplevel(self.root)
        win.title("新增/重置账号")
        win.geometry("360x300")
        frame = ttk.Frame(win, padding=18)
        frame.pack(fill="both", expand=True)
        username = StringVar()
        password = StringVar()
        role = StringVar(value="user")
        active = BooleanVar(value=True)
        ttk.Label(frame, text="账号").pack(anchor="w")
        ttk.Entry(frame, textvariable=username).pack(fill="x", pady=(3, 10))
        ttk.Label(frame, text="密码（至少 6 位）").pack(anchor="w")
        ttk.Entry(frame, textvariable=password, show="•").pack(fill="x", pady=(3, 10))
        ttk.Label(frame, text="角色").pack(anchor="w")
        ttk.Combobox(frame, textvariable=role, values=["user", "admin"], state="readonly").pack(fill="x", pady=(3, 10))
        ttk.Checkbutton(frame, text="启用账号", variable=active).pack(anchor="w", pady=(0, 12))

        def save() -> None:
            try:
                self.auth.upsert_account(username.get(), password.get(), role=role.get(), active=active.get())
            except Exception as exc:
                messagebox.showerror(APP_NAME, str(exc))
                return
            on_saved()
            win.destroy()

        ttk.Button(frame, text="保存", command=save).pack(fill="x")
        self.center_window(win)

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    setup_logging()
    logging.info("starting native app")
    try:
        ScreenshotApp().run()
    except Exception:
        logging.exception("native app crashed")
        raise


if __name__ == "__main__":
    main()
