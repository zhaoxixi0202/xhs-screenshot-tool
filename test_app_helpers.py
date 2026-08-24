import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app import content_disposition, has_uploaded_file, server_address, workbook_info_payload
from workbook import resolve_node


class AppHelperTests(unittest.TestCase):
    def test_has_uploaded_file_does_not_coerce_fieldstorage_to_bool(self):
        class BoolExplodes:
            filename = "input.xlsx"

            def __bool__(self):
                raise TypeError("Cannot be converted to bool.")

        self.assertTrue(has_uploaded_file(BoolExplodes()))

    def test_workbook_info_payload_lists_sheets_and_header_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "links.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "投放笔记"
            ws["A1"] = "达人"
            ws["B1"] = "笔记链接"
            ws["C1"] = "截图"
            wb.create_sheet("空表")
            wb.save(source)

            payload = workbook_info_payload(source)

        self.assertEqual(payload["sheets"][0]["name"], "投放笔记")
        self.assertEqual(payload["sheets"][0]["columns"][:3], [
            {"letter": "A", "label": "A - 达人"},
            {"letter": "B", "label": "B - 笔记链接"},
            {"letter": "C", "label": "C - 截图"},
        ])
        self.assertEqual(payload["recommended"]["sheetName"], "投放笔记")
        self.assertEqual(payload["recommended"]["linkCol"], "B")
        self.assertEqual(payload["recommended"]["outputCol"], "C")

    def test_content_disposition_supports_chinese_filename(self):
        header = content_disposition("测试_截图结果.xlsx")

        self.assertIn('filename="download.xlsx"', header)
        self.assertIn("filename*=UTF-8''", header)
        self.assertIn("%E6%B5%8B%E8%AF%95", header)

    def test_server_address_uses_host_and_port_environment(self):
        host, port = server_address({"HOST": "0.0.0.0", "PORT": "9090"})

        self.assertEqual(host, "0.0.0.0")
        self.assertEqual(port, 9090)

    def test_resolve_node_prefers_environment_path(self):
        self.assertEqual(resolve_node({"NODE_PATH": "/usr/bin/node"}), "/usr/bin/node")


if __name__ == "__main__":
    unittest.main()
