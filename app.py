#!/usr/bin/env python3
from __future__ import annotations

import cgi
import json
import mimetypes
import os
import threading
import time
import traceback
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from workbook import BatchOptions, process_batch, process_single


DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 8788
ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
RUNS = ROOT / "runs"
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


def has_uploaded_file(upload) -> bool:
    return upload is not None and bool(getattr(upload, "filename", ""))


def server_address(env: dict | None = None) -> tuple[str, int]:
    env = env or os.environ
    return env.get("HOST", DEFAULT_HOST), int(env.get("PORT", str(DEFAULT_PORT)))


def content_disposition(filename: str) -> str:
    quoted = quote(filename.encode("utf-8"))
    return f'attachment; filename="download.xlsx"; filename*=UTF-8\'\'{quoted}'


def workbook_info_payload(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        max_col = min(max(ws.max_column or 1, 1), 80)
        columns = []
        header_values = next(ws.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True), ())
        for index in range(1, max_col + 1):
            letter = get_column_letter(index)
            header = str(header_values[index - 1]).strip() if index <= len(header_values) and header_values[index - 1] else ""
            columns.append({"letter": letter, "label": f"{letter} - {header}" if header else letter})
        sheets.append({"name": ws.title, "columns": columns})

    first = sheets[0] if sheets else {"name": "", "columns": []}
    recommended = {
        "sheetName": first["name"],
        "linkCol": guess_column(first["columns"], ("链接", "link", "url", "笔记")),
        "outputCol": guess_column(first["columns"], ("截图", "图片", "输出", "image")),
    }
    return {"sheets": sheets, "recommended": recommended}


def guess_column(columns: list[dict], keywords: tuple[str, ...]) -> str:
    for col in columns:
        label = col["label"].lower()
        if any(keyword.lower() in label for keyword in keywords):
            return col["letter"]
    return columns[0]["letter"] if columns else "A"


def new_run_dir() -> Path:
    path = RUNS / time.strftime("%Y%m%d-%H%M%S")
    path.mkdir(parents=True, exist_ok=True)
    return path


def public_path(path: str | Path) -> str:
    p = Path(path).resolve()
    try:
        rel = str(p.relative_to(ROOT)).replace("\\", "/")
        return "/file/" + quote(rel, safe="/")
    except ValueError:
        return str(p)


def screenshot_logs_payload(runs_dir: Path = RUNS, now_struct=None) -> dict:
    now_struct = now_struct or time.localtime()
    today_prefix = time.strftime("%Y%m%d", now_struct)
    logs = []
    if not runs_dir.exists():
        return {"ok": True, "date": time.strftime("%Y-%m-%d", now_struct), "logs": []}
    for run_dir in sorted(runs_dir.iterdir(), reverse=True):
        if not run_dir.is_dir() or not run_dir.name.startswith(today_prefix):
            continue
        results_path = run_dir / "worker_results.json"
        result_data = {}
        if results_path.exists():
            try:
                result_data = json.loads(results_path.read_text(encoding="utf-8"))
            except Exception:
                result_data = {}
        results = result_data.get("results", [])
        workbook = next(run_dir.glob("*_截图结果.xlsx"), None)
        logs.append({
            "runId": run_dir.name,
            "time": f"{run_dir.name[9:11]}:{run_dir.name[11:13]}:{run_dir.name[13:15]}" if len(run_dir.name) >= 15 else run_dir.name,
            "total": len(results),
            "success": sum(1 for item in results if item.get("status") == "成功"),
            "failed": sum(1 for item in results if item.get("status") and item.get("status") != "成功"),
            "stopped": bool(result_data.get("stopped", False)),
            "reason": result_data.get("reason", ""),
            "workbook": public_path(workbook) if workbook else "",
        })
    return {"ok": True, "date": time.strftime("%Y-%m-%d", now_struct), "logs": logs}


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args) -> None:
        print("[%s] %s" % (self.log_date_time_string(), fmt % args))

    def send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            return self.serve_file(STATIC / "index.html")
        if parsed.path.startswith("/file/"):
            rel = unquote(parsed.path.removeprefix("/file/"))
            target = (ROOT / rel).resolve()
            if ROOT not in target.parents and target != ROOT:
                self.send_error(403)
                return
            return self.serve_file(target, download=True)
        if parsed.path.startswith("/api/job/"):
            return self.job_status(parsed.path.rsplit("/", 1)[-1])
        if parsed.path == "/api/logs/today":
            return self.send_json(screenshot_logs_payload())
        target = (STATIC / parsed.path.lstrip("/")).resolve()
        if STATIC in target.parents or target == STATIC:
            return self.serve_file(target)
        self.send_error(404)

    def do_HEAD(self) -> None:
        parsed = urlparse(self.path)
        download = False
        if parsed.path == "/":
            target = STATIC / "index.html"
        elif parsed.path.startswith("/file/"):
            rel = unquote(parsed.path.removeprefix("/file/"))
            target = (ROOT / rel).resolve()
            download = True
            if ROOT not in target.parents and target != ROOT:
                self.send_error(403)
                return
        else:
            target = (STATIC / parsed.path.lstrip("/")).resolve()
        if not target.exists() or not target.is_file():
            self.send_error(404)
            return
        ctype = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(target.stat().st_size))
        if download:
            self.send_header("Content-Disposition", content_disposition(target.name))
        self.end_headers()

    def serve_file(self, target: Path, download: bool = False) -> None:
        if not target.exists() or not target.is_file():
            self.send_error(404)
            return
        ctype = mimetypes.guess_type(str(target))[0] or "application/octet-stream"
        data = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        if download:
            self.send_header("Content-Disposition", content_disposition(target.name))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self) -> None:
        try:
            if self.path == "/api/single":
                return self.single()
            if self.path == "/api/workbook-info":
                return self.workbook_info()
            if self.path == "/api/batch":
                return self.batch()
            self.send_error(404)
        except Exception as exc:
            traceback.print_exc()
            self.send_json({"ok": False, "error": str(exc)}, 500)

    def read_urlencoded(self) -> dict[str, str]:
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length).decode("utf-8")
        data = parse_qs(body)
        return {k: v[0] for k, v in data.items()}

    def single(self) -> None:
        data = self.read_urlencoded()
        url = data.get("url", "").strip()
        if not url.startswith(("http://", "https://")):
            return self.send_json({"ok": False, "error": "请输入 http 或 https 链接"}, 400)
        timeout_ms = int(data.get("timeoutMs") or 25000)
        run_dir = new_run_dir()
        result = process_single(url, run_dir, timeout_ms=timeout_ms)
        first = (result.get("results") or [{}])[0]
        image = first.get("screenshot") or first.get("failureScreenshot")
        self.send_json({
            "ok": first.get("status") == "成功" and not first.get("blankish"),
            "status": first.get("status"),
            "reason": "截图疑似白屏" if first.get("blankish") else first.get("reason", ""),
            "image": public_path(image) if image else "",
            "download": public_path(image) if image else "",
        })

    def workbook_info(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": self.headers.get("Content-Type"),
        })
        upload = form["file"] if "file" in form else None
        if not has_uploaded_file(upload):
            return self.send_json({"ok": False, "error": "请上传 xlsx 文件"}, 400)
        run_dir = new_run_dir()
        input_path = run_dir / Path(upload.filename).name
        input_path.write_bytes(upload.file.read())
        payload = workbook_info_payload(input_path)
        payload["ok"] = True
        self.send_json(payload)

    def batch(self) -> None:
        form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": self.headers.get("Content-Type"),
        })
        upload = form["file"] if "file" in form else None
        if not has_uploaded_file(upload):
            return self.send_json({"ok": False, "error": "请上传 xlsx 文件"}, 400)
        run_dir = new_run_dir()
        input_path = run_dir / Path(upload.filename).name
        input_path.write_bytes(upload.file.read())
        opts = BatchOptions(
            workbook_path=input_path,
            run_dir=run_dir,
            sheet_name=form.getfirst("sheetName", "Sheet1"),
            link_col=form.getfirst("linkCol", "A"),
            output_col=form.getfirst("outputCol", "B"),
            start_row=int(form.getfirst("startRow", "2")),
            end_row=int(form.getfirst("endRow") or 0) or None,
            resume=form.getfirst("resume", "on") == "on",
            timeout_ms=int(form.getfirst("timeoutMs", "25000")),
            min_delay_ms=int(form.getfirst("minDelayMs", "3500")),
            max_delay_ms=int(form.getfirst("maxDelayMs", "60000")),
            max_retries=int(form.getfirst("maxRetries", "2")),
            max_consecutive_failures=int(form.getfirst("maxConsecutiveFailures", "5")),
        )
        job_id = uuid.uuid4().hex[:12]
        with JOBS_LOCK:
            JOBS[job_id] = {"status": "running", "runDir": str(run_dir), "startedAt": time.time(), "workbook": "", "error": ""}
        threading.Thread(target=self.run_batch_job, args=(job_id, opts), daemon=True).start()
        self.send_json({"ok": True, "jobId": job_id})

    def run_batch_job(self, job_id: str, opts: BatchOptions) -> None:
        try:
            result = process_batch(opts)
            with JOBS_LOCK:
                JOBS[job_id].update({"status": "done", "workbook": result.get("workbook", ""), "result": result})
        except Exception as exc:
            traceback.print_exc()
            with JOBS_LOCK:
                JOBS[job_id].update({"status": "error", "error": str(exc)})

    def job_status(self, job_id: str) -> None:
        with JOBS_LOCK:
            job = dict(JOBS.get(job_id) or {})
        if not job:
            return self.send_json({"ok": False, "error": "任务不存在"}, 404)
        run_dir = Path(job["runDir"])
        result_path = run_dir / "worker_results.json"
        job_path = run_dir / "worker_job.json"
        results = []
        stopped = False
        reason = job.get("error", "")
        if result_path.exists():
            try:
                data = json.loads(result_path.read_text(encoding="utf-8"))
                results = data.get("results", [])
                stopped = bool(data.get("stopped", False))
                reason = data.get("reason", reason)
            except Exception:
                pass
        total = 0
        if job_path.exists():
            try:
                total = len(json.loads(job_path.read_text(encoding="utf-8")).get("items", []))
            except Exception:
                total = 0
        if job.get("result"):
            results = job["result"].get("results", results)
            stopped = bool(job["result"].get("stopped", stopped))
            reason = job["result"].get("reason", job["result"].get("message", reason))
            total = max(total, len(results))
        success = sum(1 for r in results if r.get("status") == "成功")
        failed = len(results) - success
        workbook = job.get("workbook", "")
        self.send_json({
            "ok": True,
            "jobId": job_id,
            "status": job["status"],
            "total": total,
            "done": len(results),
            "success": success,
            "failed": failed,
            "stopped": stopped,
            "reason": reason,
            "workbook": public_path(workbook) if workbook else "",
            "results": results[-10:],
        })


def main() -> None:
    RUNS.mkdir(exist_ok=True)
    host, port = server_address()
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"小红书笔记截图工具已启动：http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
