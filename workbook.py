from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image, ImageChops, ImageStat


ROOT = Path(__file__).resolve().parent
LOCAL_NODE = Path("/Users/zhaoxixi/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node")
WORKER = ROOT / "cdp_screenshot.mjs"


def resolve_node(env: dict | None = None) -> str:
    env = env or os.environ
    if env.get("NODE_PATH"):
        return env["NODE_PATH"]
    if LOCAL_NODE.exists():
        return str(LOCAL_NODE)
    found = shutil.which("node")
    if found:
        return found
    raise RuntimeError("找不到 Node.js。请安装 node，或设置 NODE_PATH。")


@dataclass
class BatchOptions:
    workbook_path: Path
    run_dir: Path
    sheet_name: str
    link_col: str
    output_col: str
    start_row: int = 2
    end_row: int | None = None
    resume: bool = True
    timeout_ms: int = 25000
    min_delay_ms: int = 3500
    max_delay_ms: int = 60000
    max_retries: int = 2
    max_consecutive_failures: int = 5
    viewport_width: int = 1280
    viewport_height: int = 900


def column_number(value: str) -> int:
    value = str(value).strip()
    if value.isdigit():
        return int(value)
    if re.fullmatch(r"[A-Za-z]+", value):
        return column_index_from_string(value.upper())
    raise ValueError(f"无效列名：{value}")


def merged_top_left(ws, row: int, col: int) -> tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_resume_status_cols(ws, output_col: int) -> tuple[int, int, int]:
    helper_start = output_col + 1
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= output_col <= rng.max_col:
            helper_start = max(helper_start, rng.max_col + 1)
    return helper_start, helper_start + 1, helper_start + 2


def collect_items(opts: BatchOptions) -> tuple[list[dict[str, Any]], Path]:
    wb = load_workbook(opts.workbook_path)
    if opts.sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表：{opts.sheet_name}")
    ws = wb[opts.sheet_name]
    link_col = column_number(opts.link_col)
    output_col = column_number(opts.output_col)
    status_col, reason_col, path_col = find_resume_status_cols(ws, output_col)
    ws.cell(1, status_col).value = "截图状态"
    ws.cell(1, reason_col).value = "失败原因"
    ws.cell(1, path_col).value = "截图文件"

    end_row = opts.end_row or ws.max_row
    items = []
    for row in range(opts.start_row, end_row + 1):
        url = cell_text(ws.cell(row, link_col).value)
        if not url:
            continue
        target_row, target_col = merged_top_left(ws, row, output_col)
        if opts.resume and cell_text(ws.cell(row, status_col).value) == "成功":
            continue
        items.append({"index": len(items) + 1, "row": row, "targetRow": target_row, "targetCol": target_col, "url": url})

    prepared = opts.run_dir / "prepared.xlsx"
    wb.save(prepared)
    return items, prepared


def trim_black_border(image_path: Path) -> None:
    img = Image.open(image_path).convert("RGB")
    bg = Image.new("RGB", img.size, (0, 0, 0))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox:
        left, top, right, bottom = bbox
        if left or top or right != img.width or bottom != img.height:
            cropped = img.crop(bbox)
            if cropped.width > img.width * 0.75 and cropped.height > img.height * 0.75:
                cropped.save(image_path)


def is_blankish(image_path: Path) -> bool:
    img = Image.open(image_path).convert("L").resize((64, 64))
    stat = ImageStat.Stat(img)
    return stat.stddev[0] < 6 and stat.mean[0] > 235


def run_worker(items: list[dict[str, Any]], opts: BatchOptions) -> Path:
    result_path = opts.run_dir / "worker_results.json"
    job_path = opts.run_dir / "worker_job.json"
    screenshot_dir = opts.run_dir / "screenshots"
    job = {
        "items": items,
        "outputDir": str(screenshot_dir),
        "timeoutMs": opts.timeout_ms,
        "delayMs": opts.min_delay_ms,
        "minDelayMs": opts.min_delay_ms,
        "maxDelayMs": opts.max_delay_ms,
        "maxRetries": opts.max_retries,
        "maxConsecutiveFailures": opts.max_consecutive_failures,
        "viewport": {
            "width": opts.viewport_width,
            "height": opts.viewport_height,
            "deviceScaleFactor": 1,
        },
    }
    job_path.write_text(json.dumps(job, ensure_ascii=False, indent=2), encoding="utf-8")
    subprocess.run([resolve_node(), str(WORKER), "--job", str(job_path), "--out", str(result_path)], check=True)
    return result_path


def apply_results(prepared_xlsx: Path, result_path: Path, opts: BatchOptions) -> dict[str, Any]:
    data = json.loads(result_path.read_text(encoding="utf-8")) if result_path.exists() else {"results": []}
    wb = load_workbook(prepared_xlsx)
    ws = wb[opts.sheet_name]
    output_col = column_number(opts.output_col)
    status_col, reason_col, path_col = find_resume_status_cols(ws, output_col)

    for result in data.get("results", []):
        row = int(result["row"])
        ws.cell(row, status_col).value = result.get("status", "")
        ws.cell(row, reason_col).value = result.get("reason", "")
        image_path = Path(result.get("screenshot") or result.get("failureScreenshot") or "")
        if image_path.exists():
            trim_black_border(image_path)
            if result.get("status") == "成功" and is_blankish(image_path):
                ws.cell(row, status_col).value = "失败"
                ws.cell(row, reason_col).value = "截图疑似白屏"
            ws.cell(row, path_col).value = str(image_path)
            anchor_row, anchor_col = merged_top_left(ws, int(result.get("targetRow", row)), int(result.get("targetCol", output_col)))
            anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
            img = XlsxImage(str(image_path))
            img.width = 320
            img.height = 320
            ws.add_image(img, anchor)
            ws.row_dimensions[anchor_row].height = max(ws.row_dimensions[anchor_row].height or 0, 245)
            ws.column_dimensions[get_column_letter(anchor_col)].width = max(ws.column_dimensions[get_column_letter(anchor_col)].width or 0, 45)

    output = opts.run_dir / f"{opts.workbook_path.stem}_截图结果.xlsx"
    wb.save(output)
    data["workbook"] = str(output)
    result_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def process_batch(opts: BatchOptions) -> dict[str, Any]:
    opts.run_dir.mkdir(parents=True, exist_ok=True)
    items, prepared = collect_items(opts)
    if not items:
        output = opts.run_dir / f"{opts.workbook_path.stem}_截图结果.xlsx"
        Path(prepared).replace(output)
        return {"results": [], "workbook": str(output), "message": "没有需要处理的链接"}
    result_path = run_worker(items, opts)
    return apply_results(prepared, result_path, opts)


def process_single(url: str, run_dir: Path, timeout_ms: int = 25000) -> dict[str, Any]:
    run_dir.mkdir(parents=True, exist_ok=True)
    opts = BatchOptions(
        workbook_path=Path("single.xlsx"),
        run_dir=run_dir,
        sheet_name="Sheet1",
        link_col="A",
        output_col="B",
        timeout_ms=timeout_ms,
        min_delay_ms=1000,
        max_retries=1,
        max_consecutive_failures=1,
    )
    result_path = run_dir / "worker_results.json"
    screenshot_dir = run_dir / "screenshots"
    job_path = run_dir / "worker_job.json"
    item = {"index": 1, "row": 1, "targetRow": 1, "targetCol": 1, "url": url}
    job = {
        "items": [item],
        "outputDir": str(screenshot_dir),
        "timeoutMs": timeout_ms,
        "delayMs": 1000,
        "minDelayMs": 1000,
        "maxDelayMs": 5000,
        "maxRetries": 1,
        "maxConsecutiveFailures": 1,
        "viewport": {"width": opts.viewport_width, "height": opts.viewport_height, "deviceScaleFactor": 1},
    }
    job_path.write_text(json.dumps(job, ensure_ascii=False, indent=2), encoding="utf-8")
    subprocess.run([resolve_node(), str(WORKER), "--job", str(job_path), "--out", str(result_path)], check=True)
    data = json.loads(result_path.read_text(encoding="utf-8"))
    for result in data.get("results", []):
        image_path = Path(result.get("screenshot") or result.get("failureScreenshot") or "")
        if image_path.exists():
            trim_black_border(image_path)
            result["blankish"] = is_blankish(image_path)
    result_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data
